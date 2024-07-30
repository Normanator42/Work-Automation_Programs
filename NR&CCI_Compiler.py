# This program compiles all data from all inspections for every project into one excel spreadsheet, which is to undergo QA before another program transfers this data to Sydney Water

import pandas as pd
import os
import sys
import re
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Directory paths
base_dir = r"F:\NR CCI UPLOADS\PENDING"
output_file_path = os.path.join(base_dir, "compiled_data.xlsx")

warnings.filterwarnings("ignore", category=FutureWarning, message="The behavior of DataFrame concatenation with empty or all-NA entries is deprecated.")
warnings.filterwarnings("ignore", category=UserWarning, message="Parsing dates in %d/%m/%Y %H:%M:%S format when dayfirst=False")

column_headers = [
    "Attempt #", "Inspection Video(s)", "US MH", "DS MH", "Inspection Direction", "Date of inspection",
    "Time of inspection", "PackageName", "Cleaning", "Inspected Length [m]", "Pipe Asset ID", "JSA/WO",
    "Child WO", "General comment", "Section PDF Filename", "Address/Location", "Suburb", "Client Defined 2",
    "WO description", "Location Scamp", "Priority Justification", "Operational Area", "Task code"]

# Function to check for a single Excel file in the given directory
def find_excel_file(directory):
    excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx') or f.endswith('.xls')]
    if len(excel_files) != 1:
        raise Exception(f"Error: Expected exactly one Excel file in '{directory}', but found {len(excel_files)}.")
    return os.path.join(directory, excel_files[0])

# Function to extract the JSA/WO number from the parent folder name
def extract_wo_number(folder_name):
    # Try to find an 8-digit number directly first in the asset
    match = re.search(r"^(\d{8})", folder_name)
    if not match:
        # If no match, then try to find an 8-digit number after "WO", "WO ", "WO  ", "WO-", or "WO_"
        match = re.search(r"WO[\s_-]{0,2}(\d{8})", folder_name, re.IGNORECASE)
    
    if match:
        return match.group(1)
    else:
        # If no 8-digit number is found, terminate the program and print an error message
        print(f"NO WO NUMBER IN {folder_name}")
        sys.exit(1)


# Traverse directories and compile data
compiled_data = pd.DataFrame(columns=column_headers)
for folder_name in os.listdir(base_dir):
    folder_path = os.path.join(base_dir, folder_name)
    if os.path.isdir(folder_path):
        docu_path = os.path.join(folder_path, "misc", "docu")
        if os.path.exists(docu_path):
            excel_file_path = find_excel_file(docu_path)
            pdf_file_path = os.path.join(docu_path, [f for f in os.listdir(docu_path) if f.endswith('.pdf')][0])
            video_base_path = os.path.join(folder_path, "Video", "Sec")
            data = pd.read_excel(excel_file_path)
            
            data.columns = column_headers[:len(data.columns)]
            
            # Update "Section PDF Filename", "Inspection Video(s)", "PackageName", and "JSA/WO" columns
            data["Section PDF Filename"] = pdf_file_path
            data["Inspection Video(s)"] = data["Inspection Video(s)"].apply(
                lambda x: os.path.join(video_base_path, x) if pd.notna(x) else x
            )
            data["PackageName"] = data["PackageName"].fillna("Reactive")  # Set default value
            wo_number = extract_wo_number(folder_name)
            data["JSA/WO"], data["Child WO"] = wo_number, wo_number
            temp = data["PackageName"]
            temp1 = data["General comment"]
            data["General comment"] = temp
            data["PackageName"] = temp1
            compiled_data = pd.concat([compiled_data, data], ignore_index=True)

# Filter out unnecessary rows
columns_to_check = [col for col in column_headers if col != "Inspected Length [m]" and col != "Section PDF Filename" and col != "JSA/WO" and col != "PackageName" and col != "General comment"]
compiled_data = compiled_data.dropna(how='all', subset=columns_to_check)

# Sort the data by WO number, "Date of inspection", and "Time of inspection"
compiled_data = compiled_data.sort_values(by=["JSA/WO", "Date of inspection", "Time of inspection"])


# Function to remove inspections within the same JSA/WO that occur within a 10-hour span
def remove_close_inspections(df):
    df['DateTime'] = pd.to_datetime(df['Date of inspection'] + ' ' + df['Time of inspection'])
    filtered_df = pd.DataFrame(columns=df.columns)
    for wo in df['JSA/WO'].unique():
        wo_df = df[df['JSA/WO'] == wo].sort_values(by='DateTime', ascending=False)
        wo_df = wo_df.reset_index(drop=True)
        keep_rows = []
        for i in range(len(wo_df)):
            if not keep_rows:
                keep_rows.append(i)
            else:
                if (wo_df.loc[keep_rows[-1], 'DateTime'] - wo_df.loc[i, 'DateTime']).total_seconds() >= 36000:
                    keep_rows.append(i)
        filtered_df = pd.concat([filtered_df, wo_df.loc[keep_rows]], ignore_index=True)
    filtered_df = filtered_df.drop(columns=['DateTime'])
    return filtered_df

compiled_data = remove_close_inspections(compiled_data)

# Save the compiled data to an Excel file
compiled_data.to_excel(output_file_path, index=False)
print(f"Compiled data saved to '{output_file_path}'")

# Correct the headers by swapping "PackageName" and "General comment"
def correct_headers(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Find the indices of "PackageName" and "General comment"
    package_name_idx = column_headers.index("PackageName") + 1  # 1-based index
    general_comment_idx = column_headers.index("General comment") + 1  # 1-based index

    # Swap the headers
    temp = ws.cell(row=1, column=package_name_idx).value
    ws.cell(row=1, column=package_name_idx).value = ws.cell(row=1, column=general_comment_idx).value
    ws.cell(row=1, column=general_comment_idx).value = temp
    
    wb.save(file_path)
    wb.close()  # Close the workbook to release the file

# Apply formatting to the generated Excel file
def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Check if all cells in "Client Defined 2" are empty and hide the column if true
    client_defined_2_column_index = column_headers.index("Client Defined 2") + 1  # 1-based index
    client_defined_2_empty = all(ws.cell(row=row, column=client_defined_2_column_index).value is None for row in range(2, ws.max_row + 1))
    if client_defined_2_empty:
        ws.column_dimensions[get_column_letter(client_defined_2_column_index)].hidden = True

    # Set column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        # Set regular size for specific columns
        if column_headers[col[0].column - 1] in ["Inspection Video(s)"]:
            adjusted_width = 20  # Adjust this value as needed
        ws.column_dimensions[column].width = adjusted_width

    # Set the "Section PDF Filename" column to a fixed width (normal size)
    section_pdf_column_index = column_headers.index("Section PDF Filename") + 1  # 1-based index
    ws.column_dimensions[get_column_letter(section_pdf_column_index)].width = 15  # Adjust this value as needed

    # Set header font to bold
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Center align all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Color specific cells in the first row with light orange
    light_orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    columns_to_color = ["Attempt #", "General comment", "PackageName", "Cleaning", "Child WO", "WO description", "Task code", "Location Scamp", "Priority Justification", "Operational Area", "JSA/WO"]
    for col in columns_to_color:
        col_index = column_headers.index(col) + 1  # 1-based index
        ws.cell(row=1, column=col_index).fill = light_orange_fill

    wb.save(file_path)
    wb.close()  # Close the workbook to release the file

# Apply formatting
format_excel(output_file_path)

# Correct the headers
correct_headers(output_file_path)
