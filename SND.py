import openpyxl
from openpyxl import load_workbook
import datetime

# Function to get user input
def get_user_input(prompt):
    return input(prompt)

# Function to append text to a cell and ensure text wrapping is enabled
def append_to_cell(cell, text):
    if cell.value is None:
        cell.value = text
    else:
        cell.value = f"{cell.value}\n{text}"
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

# Function to capitalize comments
def capitalize_comment(comment):
    return comment.upper()

# Function to check if asset numbers exist in the spreadsheet
def check_asset_numbers(asset_list, worksheet):
    existing_assets = []
    non_existing_assets = []
    asset_column_index = 5  # Column F is the 6th column, index 5
    for asset_number in asset_list:
        found = False
        for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column):
            if row[asset_column_index].value == asset_number:
                existing_assets.append(asset_number)
                found = True
                break
        if not found:
            non_existing_assets.append(asset_number)
    return existing_assets, non_existing_assets

# Dictionary to map WO numbers to file paths
file_paths = {
    "TEST": r"F:\SND MBK TEST.xlsx",
    "91171698": r"C:\Users\JulianNorman\Dropbox (Pipe Management Aus)\Operations\ICP Operations\004_CLIENTS\1. SYDNEY WATER\18 Silt & Debris\Silt and Debris - Cyclic RC Package 2024\CYCLIC RC PACKAGES - PACKAGE 1 - WO 91171698.xlsx",
    "91171751": r"C:\Users\JulianNorman\Dropbox (Pipe Management Aus)\Operations\ICP Operations\004_CLIENTS\1. SYDNEY WATER\18 Silt & Debris\Silt and Debris - Cyclic RC Package 2024\CYCLIC RC PACKAGES - PACKAGE 2 - WO 91171751 - MOOREBANK.xlsx",
    "91171801": r"C:\Users\JulianNorman\Dropbox (Pipe Management Aus)\Operations\ICP Operations\004_CLIENTS\1. SYDNEY WATER\18 Silt & Debris\Silt and Debris - Cyclic RC Package 2024\CYCLIC RC PACKAGES - PACKAGE 3 - WO 91171801 - EAGLE VALE.xlsx",
    "91171824": r"C:\Users\JulianNorman\Dropbox (Pipe Management Aus)\Operations\ICP Operations\004_CLIENTS\1. SYDNEY WATER\18 Silt & Debris\Silt and Debris - Cyclic RC Package 2024\CYCLIC RC PACKAGES - PACKAGE 4 - WO 91171842 - LURNEA.xlsx"
}

# Get the WO number from the user
chosen_tracker = get_user_input("Select cyclic WO: 91171698, 91171751 (MOOREBANK), 91171801 (EAGLEVALE), 91171824 (LURNEA): ").strip()

# Get the file path based on user input
file_path = file_paths.get(chosen_tracker)

if file_path:
    # Load the workbook and select the active worksheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Get initial user inputs
    date_str = get_user_input("Enter the date (dd/mm/yyyy): ")
    date = datetime.datetime.strptime(date_str, "%d/%m/%Y").date()
    jds_number = get_user_input("Enter the JDS# number: ").strip()
    tc_used = get_user_input("Was TC used? (y/n): ").strip().lower()
    docket_number = get_user_input("Enter the docket number: ") if tc_used == 'y' else 'NIL'

    # Start an infinite loop to handle multiple asset entries
    while True:
        asset_numbers = get_user_input("Enter asset number(s) (space-separated) or type 'e' to end: ").strip()

        if asset_numbers.lower() == 'e':
            break

        asset_list = asset_numbers.split()

        existing_assets, non_existing_assets = check_asset_numbers(asset_list, ws)

        if non_existing_assets:
            print(f"ASSET NUMBER(S) {', '.join(non_existing_assets)} NOT FOUND IN COLUMN F.")
            continue  # Skip the rest of the loop and ask for asset numbers again

        if len(existing_assets) > 1:
            is_multi = get_user_input("Is this a multi asset? (y/n): ").strip().lower()
            has_video = get_user_input("Do these assets have a video? (y/n): ").strip().lower()
            comments = {}

            comment = get_user_input(f"Enter work comment: ").strip()
            capitalized_comment = capitalize_comment(comment)               

            for asset_number in existing_assets:
                if is_multi == 'y':
                    other_assets = [a for a in existing_assets if a != asset_number]
                    comments[asset_number] = capitalized_comment + f", MULTI WITH - {' '.join(other_assets)}"
                else:
                    metreage = get_user_input(f"{asset_number} metreage (optional): ").strip().lower()
                    if metreage:
                        comments[asset_number] = capitalized_comment + f" - {metreage}"
                    else:
                        comments[asset_number] = capitalized_comment

            is_complete = get_user_input("Is the work complete? (y/n): ").strip().lower()
            requirements = ""

            if is_complete == 'n' and is_multi == 'y':
                requirements = get_user_input("State any requirements: ").strip()

            for asset_number in existing_assets:
                for row in ws.iter_rows(min_row=1, max_col=ws.max_column):
                    if row[5].value == asset_number:
                        row[18].value = 'Y' if tc_used == 'y' else 'N'
                        row[17].value = 'Y' if has_video == 'y' else 'N'
                        append_to_cell(row[19], docket_number)
                        append_to_cell(row[14], f"{date_str} - {comments[asset_number]}")
                        append_to_cell(row[13], jds_number)
                        if is_complete == 'y':
                            row[11].value = "FINISHED"
                            row[16].value = "Complete"
                            row[12].value = date_str
                        else:
                            row[15].value = requirements if is_multi == 'y' else get_user_input(f"State requirements for {asset_number}: ").strip()
                        break

        elif len(existing_assets) == 1:
            has_video = get_user_input("Does this asset have a video? (y/n): ").strip().lower()
            comment = get_user_input("Enter work comment: ").strip()
            capitalized_comment = capitalize_comment(comment)
            is_complete = get_user_input("Is the work complete? (y/n): ").strip().lower()
            requirements = ""

            if is_complete == 'n':
                requirements = get_user_input("State any requirements: ").strip()

            for row in ws.iter_rows(min_row=1, max_col=ws.max_column):
                if row[5].value == asset_numbers:
                    row[18].value = 'Y' if tc_used == 'y' else 'N'
                    row[17].value = 'Y' if has_video == 'y' else 'N'
                    append_to_cell(row[19], docket_number)
                    append_to_cell(row[14], f"{date_str} - {capitalized_comment}")
                    append_to_cell(row[13], jds_number)
                    if is_complete == 'y':
                        row[11].value = "FINISHED"
                        row[16].value = "Complete"
                        row[12].value = date_str
                    else:
                        if requirements:
                            row[15].value = requirements
                    break

    # Save the workbook
    wb.save(file_path)
    print("Workbook updated successfully.")
else:
    print("INVALID WO NUMBER SELECTED.")
